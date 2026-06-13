# ===
# File Summary
# Path: backend\apps\imports\views.py
# Type: python
# Purpose: Imports domain for parser/mapping workflows and linked entity updates.
# Primary responsibilities:
# - Domain behavior is summarized for fast onboarding and avoids full-file reread.
# - Core symbols: IsAuthenticated, has_permission, ImportJobListCreateView, post, ImportJobDryRunView
# Inputs:
# - Downstream and upstream interactions in the same domain.
# Outputs:
# - API payloads, records, side effects, or UI views depending on file role.
# Dependencies:
# - Shared runtime services and adjacent domain modules.
# Known risks:
# - Validate behavior after migrations, dependency upgrades, or contract changes.
# ===
# 

from io import BytesIO
import json

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from rest_framework import permissions, status
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.permissions import user_can
from apps.audit.services import record_audit_event
from apps.folders.models import FolderChangeEvent
from apps.imports.models import ImportAuditEvent, ImportJob
from apps.imports.services import (
    accept_folder_links,
    apply_import,
    dry_run_import,
    scan_legacy_folders,
    visible_records_for_export,
)
from apps.imports.parsers import XlsxParseError, parse_xlsx_rows
from apps.projects.models import Project, ProjectEvent
from apps.records.validation import get_object_type_definition
from apps.workflows.models import WorkflowEvent


class IsAuthenticated(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.is_active)


class ImportJobListCreateView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        object_type_key = request.data.get("target_object_type")
        if not object_type_key:
            raise ValidationError({"target_object_type": ["This field is required."]})
        if not user_can(request.user, "create", object_type_key):
            raise PermissionDenied("You do not have permission to import this object type.")
        source_file = request.FILES.get("source_file")
        if not source_file:
            raise ValidationError({"source_file": ["This field is required."]})
        job = ImportJob.objects.create(
            source_file=source_file,
            target_object_type=object_type_key,
            mapping=_json_object(request.data.get("mapping", {}), "mapping"),
            created_by=request.user,
            updated_by=request.user,
        )
        record_audit_event(
            request.user,
            "import.created",
            job,
            before=None,
            after=_job_snapshot(job),
            request=request,
        )
        return Response(_serialize_job(job), status=status.HTTP_201_CREATED)


class ImportJobDryRunView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        job = get_object_or_404(ImportJob, pk=pk)
        if not _can_import_or_edit(request.user, job.target_object_type):
            raise PermissionDenied("You do not have permission to import this object type.")
        before = _job_snapshot(job)
        try:
            result = dry_run_import(job, actor=request.user)
        except Exception:
            job.refresh_from_db()
            record_audit_event(
                request.user,
                "import.dry_run",
                job,
                before=before,
                after=_job_snapshot(job),
                request=request,
            )
            raise
        job.refresh_from_db()
        record_audit_event(
            request.user,
            "import.dry_run",
            job,
            before=before,
            after=_job_snapshot(job),
            request=request,
        )
        return Response(result, status=status.HTTP_200_OK)


class ImportJobApplyView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        job = get_object_or_404(ImportJob, pk=pk)
        create_folders = bool(request.data.get("create_managed_folders", False))
        before = _job_snapshot(job)
        result = apply_import(
            job,
            actor=request.user,
            create_managed_folders=create_folders,
            request=request,
        )
        job.refresh_from_db()
        record_audit_event(
            request.user,
            "import.applied",
            job,
            before=before,
            after=_job_snapshot(job),
            request=request,
        )
        return Response(result, status=status.HTTP_200_OK)


class ImportColumnPreviewView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        source_file = request.FILES.get("source_file")
        if not source_file:
            raise ValidationError({"source_file": ["This field is required."]})

        try:
            rows = parse_xlsx_rows(source_file)
        except XlsxParseError as error:
            raise ValidationError({"source_file": [str(error)]}) from error

        first_row = rows[0] if rows else {}
        columns = list(first_row.get("values", {}).keys()) if isinstance(first_row, dict) else []
        return Response({"columns": columns}, status=status.HTTP_200_OK)


class FolderScanView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        object_type_key = request.data.get("object_type_key")
        if not user_can(request.user, "admin", object_type_key):
            raise PermissionDenied("You do not have permission to link folders for this object type.")
        result = scan_legacy_folders(
            legacy_root_path=request.data.get("legacy_root_path"),
            object_type_key=object_type_key,
            matching_rule=_json_object(
                request.data.get("matching_rule", {"type": "code_in_path"}),
                "matching_rule",
            ),
        )
        return Response(result, status=status.HTTP_200_OK)


class FolderLinkAcceptView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        result = accept_folder_links(
            legacy_root_path=request.data.get("legacy_root_path"),
            object_type_key=request.data.get("object_type_key"),
            links=request.data.get("links", []),
            actor=request.user,
            request=request,
        )
        return Response(result, status=status.HTTP_200_OK)


class RecordsExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, object_type_key):
        records = visible_records_for_export(request.user, object_type_key)
        object_type, _active_config = get_object_type_definition(object_type_key)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = object_type.get("plural_label") or object_type_key
        fields = object_type.get("fields", [])
        sheet.append(["Code", "Title", "Status", *[field.get("label", field["key"]) for field in fields]])
        for record in records:
            sheet.append(
                [
                    record.code,
                    record.title,
                    record.status,
                    *[_excel_value(record.data.get(field["key"])) for field in fields],
                ]
            )
        return _xlsx_response(workbook, f"{object_type_key}.xlsx")


class AuditExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Audit"
        sheet.append(["Source", "Action", "Created At", "Record Code", "Actor", "Details"])
        _append_import_events(sheet, request.user)
        _append_folder_events(sheet, request.user)
        _append_project_events(sheet, request.user)
        _append_workflow_events(sheet, request.user)
        return _xlsx_response(workbook, "audit.xlsx")


class ProjectStatusExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Project Status"
        sheet.append(["Code", "Project", "Status", "Open Tasks", "Total Tasks"])
        projects = Project.objects.select_related("record").prefetch_related("tasks").order_by("name")
        appended_rows = 0
        for project in projects:
            if not user_can(request.user, "view", "project", record_id=str(project.record_id)):
                continue
            tasks = list(project.tasks.all())
            open_tasks = [task for task in tasks if task.state != "done"]
            sheet.append(
                [
                    project.record.code,
                    project.name,
                    project.status,
                    len(open_tasks),
                    len(tasks),
                ]
            )
            appended_rows += 1
        if appended_rows == 0 and not user_can(request.user, "view", "project"):
            raise PermissionDenied("You do not have permission to view projects.")
        return _xlsx_response(workbook, "project-status.xlsx")


def _serialize_job(job):
    return {
        "id": job.pk,
        "target_object_type": job.target_object_type,
        "mapping": job.mapping,
        "dry_run_results": job.dry_run_results,
        "created_records_count": job.created_records_count,
        "updated_records_count": job.updated_records_count,
        "error_rows": job.error_rows,
        "state": job.state,
    }


def _job_snapshot(job):
    return {
        "id": job.pk,
        "target_object_type": job.target_object_type,
        "mapping": job.mapping,
        "dry_run_results": job.dry_run_results,
        "created_records_count": job.created_records_count,
        "updated_records_count": job.updated_records_count,
        "error_rows": job.error_rows,
        "state": job.state,
    }


def _json_object(value, field_name):
    if isinstance(value, dict):
        return value
    if value is None or value == "":
        return {}
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError as error:
            raise ValidationError({field_name: ["Expected a JSON object."]}) from error
        if isinstance(parsed, dict):
            return parsed
    raise ValidationError({field_name: ["Expected a JSON object."]})


def _can_import_or_edit(user, object_type_key):
    return (
        user_can(user, "create", object_type_key)
        or user_can(user, "edit", object_type_key)
        or user_can(user, "admin", object_type_key)
    )


def _append_import_events(sheet, user):
    events = ImportAuditEvent.objects.select_related("record", "actor").order_by("created_at", "id")
    for event in events:
        if event.record and not user_can(
            user,
            "view",
            event.record.object_type_key,
            record_id=str(event.record_id),
        ):
            continue
        sheet.append(
            [
                "import",
                event.action,
                event.created_at.isoformat(),
                event.record.code if event.record else "",
                event.actor.username if event.actor else "",
                str(event.data),
            ]
        )


def _append_folder_events(sheet, user):
    events = FolderChangeEvent.objects.select_related("matched_record", "reviewer").order_by(
        "created_at",
        "id",
    )
    for event in events:
        record = event.matched_record
        if record and not user_can(user, "view", record.object_type_key, record_id=str(record.pk)):
            continue
        sheet.append(
            [
                "folder",
                event.event_type,
                event.created_at.isoformat(),
                record.code if record else "",
                event.reviewer.username if event.reviewer else "",
                event.path,
            ]
        )


def _append_project_events(sheet, user):
    events = ProjectEvent.objects.select_related("project__record", "actor").order_by(
        "created_at",
        "id",
    )
    for event in events:
        if not user_can(user, "view", "project", record_id=str(event.project.record_id)):
            continue
        sheet.append(
            [
                "project",
                event.action,
                event.created_at.isoformat(),
                event.project.record.code,
                event.actor.username if event.actor else "",
                event.comment or str(event.data),
            ]
        )


def _append_workflow_events(sheet, user):
    events = WorkflowEvent.objects.select_related("instance__record", "actor").order_by(
        "created_at",
        "id",
    )
    for event in events:
        record = event.instance.record
        if not user_can(user, "view", record.object_type_key, record_id=str(record.pk)):
            continue
        sheet.append(
            [
                "workflow",
                event.action,
                event.created_at.isoformat(),
                record.code,
                event.actor.username if event.actor else "",
                event.comment or str(event.data),
            ]
        )


def _xlsx_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _excel_value(value):
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, bool | int | float) or value is None:
        return value
    return str(value)

