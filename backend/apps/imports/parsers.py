# ===
# File Summary
# Path: backend\apps\imports\parsers.py
# Type: python
# Purpose: Imports domain for parser/mapping workflows and linked entity updates.
# Primary responsibilities:
# - Domain behavior is summarized for fast onboarding and avoids full-file reread.
# - Core symbols: XlsxParseError, parse_xlsx_rows, _parse_csv_rows, _is_csv_file, _cell_to_value
# Inputs:
# - Downstream and upstream interactions in the same domain.
# Outputs:
# - API payloads, records, side effects, or UI views depending on file role.
# Dependencies:
# - Shared runtime services and adjacent domain modules.
# Known risks:
# - Validate behavior after migrations, dependency upgrades, or contract changes.
# ===
# 

import csv
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class XlsxParseError(ValueError):
    pass


def parse_xlsx_rows(source_file):
    if not source_file:
        raise XlsxParseError("Import file must be a valid .xlsx workbook.")

    if _is_csv_file(source_file):
        return _parse_csv_rows(source_file)

    source_file.open("rb")
    try:
        try:
            workbook = load_workbook(source_file, data_only=True)
        except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
            raise XlsxParseError("Import file must be a valid .xlsx workbook.") from error
    finally:
        source_file.close()
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise XlsxParseError("Import file must contain a header row.")

    headers = [_cell_to_value(value) for value in rows[0]]
    if not any(headers):
        raise XlsxParseError("Import file must contain a header row.")
    parsed = []
    for index, row in enumerate(rows[1:], start=2):
        if all(value is None or value == "" for value in row):
            continue
        parsed.append(
            {
                "row_number": index,
                "values": {
                    header: _cell_to_value(row[column_index])
                    for column_index, header in enumerate(headers)
                    if header
                },
            }
        )
    return parsed


def _parse_csv_rows(source_file):
    source_file.open("rb")
    try:
        content = source_file.read().decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise XlsxParseError("Import file must be a valid UTF-8 .csv file.") from error
    finally:
        source_file.close()

    reader = csv.DictReader(content.splitlines())
    headers = [header for header in (reader.fieldnames or []) if header]
    if not headers:
        raise XlsxParseError("Import file must contain a header row.")

    parsed = []
    for index, row in enumerate(reader, start=2):
        values = {header: _cell_to_value(row.get(header)) for header in headers}
        if all(value is None or value == "" for value in values.values()):
            continue
        parsed.append({"row_number": index, "values": values})
    return parsed


def _is_csv_file(source_file):
    return Path(getattr(source_file, "name", "")).suffix.lower() == ".csv"


def _cell_to_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value

