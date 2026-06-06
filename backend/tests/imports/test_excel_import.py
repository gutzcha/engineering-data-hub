from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.base import ContentFile
from openpyxl import Workbook, load_workbook

from apps.accounts.models import ObjectPermission
from apps.config_registry.services import create_draft_from_current, publish_draft
from apps.records.models import Record


@pytest.fixture
def user_factory(db):
    User = get_user_model()

    def create_user(username, role_name=None, *, is_superuser=False):
        user = User.objects.create_user(
            username=username,
            password="test-pass",
            is_superuser=is_superuser,
        )
        if role_name:
            group, _created = Group.objects.get_or_create(name=role_name)
            user.groups.add(group)
        return user

    return create_user


@pytest.fixture
def active_import_config(user_factory):
    publisher = user_factory("import-config-publisher")
    draft = create_draft_from_current(publisher)
    draft.data = {
        "object_types": [
            {
                "key": "product",
                "label": "Product",
                "plural_label": "Products",
                "code_pattern": "PROD-{seq:000000}",
                "title_field": "commercial_name",
                "fields": [
                    {
                        "key": "commercial_name",
                        "label": "Commercial Name",
                        "type": "text",
                        "required": True,
                        "unique": True,
                    },
                    {
                        "key": "category",
                        "label": "Category",
                        "type": "choice",
                        "options": ["film", "resin"],
                    },
                ],
            }
        ],
        "relationship_types": [],
        "form_layouts": [],
        "folder_templates": [],
        "dashboards": [],
    }
    draft.save()
    return publish_draft(draft, publisher)


@pytest.fixture
def import_permissions(db):
    ObjectPermission.objects.create(
        role_name="Importer",
        object_type_key="product",
        can_view=True,
        can_create=True,
        can_edit=True,
        can_admin=True,
    )
    ObjectPermission.objects.create(
        role_name="Creator",
        object_type_key="product",
        can_view=True,
        can_create=True,
    )
    ObjectPermission.objects.create(
        role_name="Editor",
        object_type_key="product",
        can_view=True,
        can_create=True,
        can_edit=True,
    )
    ObjectPermission.objects.create(
        role_name="Viewer",
        object_type_key="product",
        can_view=True,
    )
    ObjectPermission.objects.create(
        role_name="Viewer",
        object_type_key="project",
        can_view=True,
    )


def workbook_file(rows, name="records.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return ContentFile(output.getvalue(), name=name)


def import_mapping():
    return {
        "columns": {
            "Code": "code",
            "Commercial Name": "commercial_name",
            "Category": "category",
        }
    }


@pytest.mark.django_db
def test_excel_dry_run_reports_creates_updates_and_row_errors(
    active_import_config,
    import_permissions,
):
    from apps.imports.models import ImportJob
    from apps.imports.services import dry_run_import

    Record.objects.create(
        object_type_key="product",
        code="PROD-000001",
        title="Existing Film",
        schema_version=active_import_config.version,
        data={"commercial_name": "Existing Film", "category": "film"},
    )
    job = ImportJob.objects.create(
        source_file=workbook_file(
            [
                ["Code", "Commercial Name", "Category"],
                ["PROD-000002", "New Film", "film"],
                ["PROD-000001", "Updated Film", "resin"],
                ["PROD-000003", "", "film"],
                ["PROD-000004", "Bad Choice", "liquid"],
                ["PROD-000005", "Duplicate A", "film"],
                ["PROD-000005", "Duplicate B", "resin"],
            ]
        ),
        target_object_type="product",
        mapping=import_mapping(),
    )

    result = dry_run_import(job)

    job.refresh_from_db()
    assert job.state == ImportJob.State.DRY_RUN_FAILED
    assert result["summary"] == {"create": 1, "update": 1, "errors": 4}
    assert [row["code"] for row in result["creates"]] == ["PROD-000002"]
    assert [row["code"] for row in result["updates"]] == ["PROD-000001"]
    errors_by_row = {error["row_number"]: error for error in result["error_rows"]}
    assert errors_by_row[4]["errors"]["commercial_name"] == ["This field is required."]
    assert errors_by_row[5]["errors"]["category"] == ["Value must be one of: film, resin."]
    assert errors_by_row[6]["errors"]["code"] == ["Duplicate code in import file."]
    assert errors_by_row[7]["errors"]["code"] == ["Duplicate code in import file."]


@pytest.mark.django_db
def test_apply_requires_clean_dry_run_then_writes_records_events_and_folders(
    active_import_config,
    import_permissions,
    settings,
    tmp_path,
    user_factory,
    django_capture_on_commit_callbacks,
):
    from apps.folders.models import ManagedFolder
    from apps.imports.models import ImportAuditEvent, ImportJob
    from apps.imports.services import apply_import, dry_run_import

    settings.MANAGED_FILE_ROOT = tmp_path
    actor = user_factory("record-importer", "Importer")
    existing = Record.objects.create(
        object_type_key="product",
        code="PROD-000001",
        title="Existing Film",
        schema_version=active_import_config.version,
        data={"commercial_name": "Existing Film", "category": "film"},
    )
    blocked_job = ImportJob.objects.create(
        source_file=workbook_file(
            [["Code", "Commercial Name", "Category"], ["PROD-000099", "", "film"]],
            name="blocked.xlsx",
        ),
        target_object_type="product",
        mapping=import_mapping(),
    )
    dry_run_import(blocked_job)

    with pytest.raises(ValueError, match="Dry-run must pass"):
        apply_import(blocked_job, actor=actor)

    job = ImportJob.objects.create(
        source_file=workbook_file(
            [
                ["Code", "Commercial Name", "Category"],
                ["PROD-000001", "Updated Film", "resin"],
                ["PROD-000002", "New Film", "film"],
            ],
            name="clean.xlsx",
        ),
        target_object_type="product",
        mapping=import_mapping(),
    )
    dry_run_import(job)

    with django_capture_on_commit_callbacks(execute=True):
        result = apply_import(job, actor=actor, create_managed_folders=True)

    job.refresh_from_db()
    existing.refresh_from_db()
    created = Record.objects.get(code="PROD-000002")
    assert job.state == ImportJob.State.APPLIED
    assert result == {"created": 1, "updated": 1}
    assert existing.title == "Updated Film"
    assert existing.data["category"] == "resin"
    assert created.title == "New Film"
    assert ManagedFolder.objects.filter(record=created, folder_role="primary").exists()
    assert ImportAuditEvent.objects.filter(
        job=job,
        record=existing,
        action=ImportAuditEvent.Action.UPDATED,
        actor=actor,
    ).exists()
    assert ImportAuditEvent.objects.filter(
        job=job,
        record=created,
        action=ImportAuditEvent.Action.CREATED,
        actor=actor,
    ).exists()


@pytest.mark.django_db
def test_non_admin_cannot_create_manual_codes_but_can_update_by_code(
    active_import_config,
    import_permissions,
    user_factory,
):
    from apps.imports.models import ImportJob
    from apps.imports.services import apply_import, dry_run_import

    actor = user_factory("non-admin-editor", "Editor")
    existing = Record.objects.create(
        object_type_key="product",
        code="PROD-000001",
        title="Existing Film",
        schema_version=active_import_config.version,
        data={"commercial_name": "Existing Film", "category": "film"},
    )
    manual_create_job = ImportJob.objects.create(
        source_file=workbook_file(
            [["Code", "Commercial Name", "Category"], ["PROD-000777", "Manual Film", "film"]]
        ),
        target_object_type="product",
        mapping=import_mapping(),
    )

    manual_result = dry_run_import(manual_create_job, actor=actor)

    assert manual_create_job.state == ImportJob.State.DRY_RUN_FAILED
    assert manual_result["error_rows"][0]["errors"]["code"] == [
        "Manual codes for new records require admin permission."
    ]
    with pytest.raises(ValueError, match="Dry-run must pass"):
        apply_import(manual_create_job, actor=actor)

    update_job = ImportJob.objects.create(
        source_file=workbook_file(
            [["Code", "Commercial Name", "Category"], ["PROD-000001", "Updated Film", "resin"]],
            name="update.xlsx",
        ),
        target_object_type="product",
        mapping=import_mapping(),
    )

    update_result = dry_run_import(update_job, actor=actor)
    apply_result = apply_import(update_job, actor=actor)

    existing.refresh_from_db()
    assert update_result["summary"] == {"create": 0, "update": 1, "errors": 0}
    assert apply_result == {"created": 0, "updated": 1}
    assert existing.title == "Updated Film"


@pytest.mark.django_db
def test_apply_preflight_prevents_orphan_folders_when_later_row_would_fail(
    active_import_config,
    import_permissions,
    settings,
    tmp_path,
    user_factory,
):
    from apps.imports.models import ImportJob
    from apps.imports.services import apply_import, dry_run_import

    settings.MANAGED_FILE_ROOT = tmp_path
    actor = user_factory("create-only-importer", "Creator")
    Record.objects.create(
        object_type_key="product",
        code="PROD-000010",
        title="Existing Film",
        schema_version=active_import_config.version,
        data={"commercial_name": "Existing Film", "category": "film"},
    )
    job = ImportJob.objects.create(
        source_file=workbook_file(
            [
                ["Code", "Commercial Name", "Category"],
                ["", "New Auto Film", "film"],
                ["PROD-000010", "Blocked Update", "resin"],
            ],
            name="orphan-preflight.xlsx",
        ),
        target_object_type="product",
        mapping=import_mapping(),
    )
    dry_run_import(job)

    with pytest.raises(ValueError, match="Dry-run must pass"):
        apply_import(job, actor=actor, create_managed_folders=True)

    assert not (tmp_path / "Products").exists()
    assert Record.objects.filter(title="New Auto Film").exists() is False


@pytest.mark.django_db
def test_dry_run_catches_global_code_collision_and_intrafile_unique_duplicates(
    active_import_config,
    import_permissions,
):
    from apps.imports.models import ImportJob
    from apps.imports.services import dry_run_import

    Record.objects.create(
        object_type_key="raw_material",
        code="PROD-999999",
        title="Other Type Collision",
        schema_version=active_import_config.version,
        data={},
    )
    job = ImportJob.objects.create(
        source_file=workbook_file(
            [
                ["Code", "Commercial Name", "Category"],
                ["PROD-999999", "Global Collision", "film"],
                ["PROD-000100", "Repeated Commercial", "film"],
                ["PROD-000101", "Repeated Commercial", "resin"],
            ],
            name="collisions.xlsx",
        ),
        target_object_type="product",
        mapping=import_mapping(),
    )

    result = dry_run_import(job)

    assert job.state == ImportJob.State.DRY_RUN_FAILED
    errors_by_row = {error["row_number"]: error for error in result["error_rows"]}
    assert errors_by_row[2]["errors"]["code"] == [
        "Record code already exists on another object type."
    ]
    assert errors_by_row[3]["errors"]["commercial_name"] == [
        "Duplicate unique value in import file."
    ]
    assert errors_by_row[4]["errors"]["commercial_name"] == [
        "Duplicate unique value in import file."
    ]


@pytest.mark.django_db
def test_malformed_workbook_dry_run_route_returns_validation_error(
    client,
    active_import_config,
    import_permissions,
    user_factory,
):
    from apps.imports.models import ImportJob

    actor = user_factory("malformed-importer", "Importer")
    job = ImportJob.objects.create(
        source_file=ContentFile(b"not an xlsx file", name="bad.xlsx"),
        target_object_type="product",
        mapping=import_mapping(),
    )
    client.force_login(actor)
    client.raise_request_exception = False

    response = client.post(f"/api/imports/jobs/{job.pk}/dry-run/")

    job.refresh_from_db()
    assert response.status_code == 400
    assert response.json()["source_file"] == ["Import file must be a valid .xlsx workbook."]
    assert job.state == ImportJob.State.DRY_RUN_FAILED


@pytest.mark.django_db
def test_export_routes_return_visible_records_audit_and_project_status_workbooks(
    client,
    active_import_config,
    import_permissions,
    user_factory,
):
    from apps.imports.models import ImportAuditEvent, ImportJob
    from apps.projects.models import Project, ProjectTask

    actor = user_factory("export-viewer", "Viewer")
    record = Record.objects.create(
        object_type_key="product",
        code="PROD-000010",
        title="Exported Film",
        schema_version=active_import_config.version,
        data={"commercial_name": "Exported Film", "category": "film"},
    )
    job = ImportJob.objects.create(target_object_type="product", mapping=import_mapping())
    ImportAuditEvent.objects.create(
        job=job,
        record=record,
        action=ImportAuditEvent.Action.CREATED,
        actor=actor,
        data={"code": record.code},
    )
    project_record = Record.objects.create(
        object_type_key="project",
        code="PRJ-000001",
        title="Status Project",
        schema_version=1,
        data={"project_name": "Status Project"},
    )
    project = Project.objects.create(record=project_record, name="Status Project", status="active")
    ProjectTask.objects.create(project=project, title="Open task", state="in_progress")
    client.force_login(actor)

    records_response = client.get("/api/exports/records/product.xlsx")
    audit_response = client.get("/api/exports/audit.xlsx")
    project_response = client.get("/api/exports/project-status.xlsx")

    assert records_response.status_code == 200
    records_workbook = load_workbook(BytesIO(records_response.content))
    records_sheet = records_workbook.active
    assert [cell.value for cell in records_sheet[1]] == [
        "Code",
        "Title",
        "Status",
        "Commercial Name",
        "Category",
    ]
    assert [cell.value for cell in records_sheet[2]] == [
        "PROD-000010",
        "Exported Film",
        "draft",
        "Exported Film",
        "film",
    ]

    assert audit_response.status_code == 200
    audit_rows = list(load_workbook(BytesIO(audit_response.content)).active.iter_rows(values_only=True))
    assert ("import", "created", "PROD-000010", "export-viewer") in [
        (row[0], row[1], row[3], row[4]) for row in audit_rows[1:]
    ]

    assert project_response.status_code == 200
    project_rows = list(load_workbook(BytesIO(project_response.content)).active.iter_rows(values_only=True))
    assert project_rows[1][:5] == ("PRJ-000001", "Status Project", "active", 1, 1)
